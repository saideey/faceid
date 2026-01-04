from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date
from calendar import monthrange
from database import get_db, Employee, AttendanceLog, Penalty, Department
from sqlalchemy import func, and_, extract
from decimal import Decimal
import os
from config.settings import Config
from database import CompanyAdmin
from sqlalchemy import Boolean


def generate_monthly_excel(company_id, year, month):
    """Generate monthly attendance report in Excel format"""
    db = get_db()
    try:
        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        create_summary_sheet(wb, db, company_id, year, month)
        create_employee_breakdown_sheet(wb, db, company_id, year, month)
        create_daily_attendance_sheet(wb, db, company_id, year, month)

        # Generate filename
        filename = f"attendance_report_{company_id}_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(Config.EXPORT_FOLDER, filename)

        # Save workbook
        wb.save(filepath)

        return filename, None
    except Exception as e:
        return None, str(e)
    finally:
        db.close()


def create_summary_sheet(wb, db, company_id, year, month):
    """Create summary statistics sheet"""
    ws = wb.create_sheet("Summary")

    # Get statistics
    stats = get_monthly_statistics(company_id, year, month)

    # Header
    ws['A1'] = 'Monthly Attendance Summary'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')

    # Period
    ws['A2'] = 'Period:'
    ws['B2'] = f"{year}-{month:02d}"

    # Statistics
    row = 4
    stats_data = [
        ('Total Employees', stats.get('total_employees', 0)),
        ('Total Working Days', stats.get('total_working_days', 0)),
        ('Total Present', stats.get('total_present', 0)),
        ('Total Absent', stats.get('total_absent', 0)),
        ('Total Late Arrivals', stats.get('total_late', 0)),
        ('Total Early Leaves', stats.get('total_early_leaves', 0)),
        ('Total Penalties Amount', f"{stats.get('total_penalties', 0)} UZS"),
        ('Average Work Hours', f"{stats.get('avg_work_hours', 0):.2f} hours"),
    ]

    for label, value in stats_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def create_employee_breakdown_sheet(wb, db, company_id, year, month):
    """Create employee-wise breakdown sheet"""
    ws = wb.create_sheet("Employee Breakdown")

    # Headers
    headers = ['Employee No', 'Name', 'Department', 'Days Present', 'Days Late',
               'Total Late Minutes', 'Total Penalties', 'Avg Work Hours']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Get employee data
    start_date = date(year, month, 1)
    _, last_day = monthrange(year, month)
    end_date = date(year, month, last_day)

    employees = db.query(Employee).filter_by(company_id=company_id, status='active').all()

    row = 2
    for employee in employees:
        # Get attendance logs
        logs = db.query(AttendanceLog).filter(
            and_(
                AttendanceLog.employee_id == employee.id,
                AttendanceLog.date >= start_date,
                AttendanceLog.date <= end_date
            )
        ).all()

        # Get penalties
        penalties = db.query(Penalty).filter(
            and_(
                Penalty.employee_id == employee.id,
                Penalty.date >= start_date,
                Penalty.date <= end_date
            )
        ).all()

        days_present = len(logs)
        days_late = len([l for l in logs if l.late_minutes > 0])
        total_late_minutes = sum(l.late_minutes for l in logs)
        total_penalties = sum(p.amount for p in penalties)
        total_work_minutes = sum(l.total_work_minutes for l in logs if l.total_work_minutes)
        avg_work_hours = (total_work_minutes / 60 / days_present) if days_present > 0 else 0

        # Get department name
        dept_name = employee.department.name if employee.department else 'N/A'

        ws.cell(row=row, column=1, value=employee.employee_no)
        ws.cell(row=row, column=2, value=employee.full_name)
        ws.cell(row=row, column=3, value=dept_name)
        ws.cell(row=row, column=4, value=days_present)
        ws.cell(row=row, column=5, value=days_late)
        ws.cell(row=row, column=6, value=total_late_minutes)
        ws.cell(row=row, column=7, value=f"{float(total_penalties):.2f}")
        ws.cell(row=row, column=8, value=f"{avg_work_hours:.2f}")

        row += 1

    # Format columns
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15


def create_daily_attendance_sheet(wb, db, company_id, year, month):
    """Create daily attendance sheet"""
    ws = wb.create_sheet("Daily Attendance")

    # Headers
    headers = ['Date', 'Employee No', 'Name', 'Check In', 'Check Out',
               'Late (min)', 'Work Hours', 'Status']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Get attendance logs
    start_date = date(year, month, 1)
    _, last_day = monthrange(year, month)
    end_date = date(year, month, last_day)

    logs = db.query(AttendanceLog).join(Employee).filter(
        and_(
            AttendanceLog.company_id == company_id,
            AttendanceLog.date >= start_date,
            AttendanceLog.date <= end_date
        )
    ).order_by(AttendanceLog.date.desc(), Employee.employee_no).all()

    row = 2
    for log in logs:
        work_hours = (log.total_work_minutes / 60) if log.total_work_minutes else 0
        status = 'On Time' if log.late_minutes == 0 else f'Late ({log.late_minutes} min)'

        ws.cell(row=row, column=1, value=log.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=log.employee_no)
        ws.cell(row=row, column=3, value=log.employee.full_name)
        ws.cell(row=row, column=4, value=log.check_in_time.strftime('%H:%M:%S') if log.check_in_time else '')
        ws.cell(row=row, column=5, value=log.check_out_time.strftime('%H:%M:%S') if log.check_out_time else '')
        ws.cell(row=row, column=6, value=log.late_minutes)
        ws.cell(row=row, column=7, value=f"{work_hours:.2f}")
        ws.cell(row=row, column=8, value=status)

        row += 1

    # Format columns
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15


def get_daily_statistics(company_id, target_date):
    """Get statistics for a specific date"""
    db = get_db()
    try:
        # Total employees
        total_employees = db.query(Employee).filter_by(
            company_id=company_id,
            status='active'
        ).count()

        # Attendance logs for the date
        logs = db.query(AttendanceLog).filter(
            and_(
                AttendanceLog.company_id == company_id,
                AttendanceLog.date == target_date
            )
        ).all()

        present_count = len(logs)
        late_count = len([l for l in logs if l.late_minutes > 0])
        on_time_count = present_count - late_count
        absent_count = total_employees - present_count

        # Penalties for the date
        total_penalties = db.query(func.sum(Penalty.amount)).filter(
            and_(
                Penalty.company_id == company_id,
                Penalty.date == target_date
            )
        ).scalar() or Decimal('0.00')

        return {
            'date': target_date.isoformat(),
            'total_employees': total_employees,
            'present': present_count,
            'absent': absent_count,
            'on_time': on_time_count,
            'late': late_count,
            'total_penalties': float(total_penalties)
        }, None
    except Exception as e:
        return None, str(e)
    finally:
        db.close()


def get_monthly_statistics(company_id, year, month):
    """Get statistics for a month"""
    db = get_db()
    try:
        start_date = date(year, month, 1)
        _, last_day = monthrange(year, month)
        end_date = date(year, month, last_day)

        # Total employees
        total_employees = db.query(Employee).filter_by(
            company_id=company_id,
            status='active'
        ).count()

        # Attendance logs for the month
        logs = db.query(AttendanceLog).filter(
            and_(
                AttendanceLog.company_id == company_id,
                AttendanceLog.date >= start_date,
                AttendanceLog.date <= end_date
            )
        ).all()

        # Calculate statistics
        total_present = len(logs)
        total_late = len([l for l in logs if l.late_minutes > 0])
        total_early_leaves = len([l for l in logs if l.early_leave_minutes > 0])

        # Working days (simplified - actual days in month)
        total_working_days = last_day

        # Total absent (employees * working days - present)
        total_absent = (total_employees * total_working_days) - total_present

        # Total penalties
        total_penalties = db.query(func.sum(Penalty.amount)).filter(
            and_(
                Penalty.company_id == company_id,
                Penalty.date >= start_date,
                Penalty.date <= end_date
            )
        ).scalar() or Decimal('0.00')

        # Average work hours
        total_work_minutes = sum(l.total_work_minutes for l in logs if l.total_work_minutes)
        avg_work_hours = (total_work_minutes / 60 / total_present) if total_present > 0 else 0

        return {
            'total_employees': total_employees,
            'total_working_days': total_working_days,
            'total_present': total_present,
            'total_absent': total_absent,
            'total_late': total_late,
            'total_early_leaves': total_early_leaves,
            'total_penalties': float(total_penalties),
            'avg_work_hours': avg_work_hours
        }
    except Exception as e:
        return {}
    finally:
        db.close()


def get_employee_summary(employee_id, start_date=None, end_date=None):
    """Get summary statistics for a specific employee"""
    db = get_db()
    try:
        query = db.query(AttendanceLog).filter_by(employee_id=employee_id)

        if start_date:
            query = query.filter(AttendanceLog.date >= start_date)
        if end_date:
            query = query.filter(AttendanceLog.date <= end_date)

        logs = query.all()

        days_present = len(logs)
        days_late = len([l for l in logs if l.late_minutes > 0])
        total_late_minutes = sum(l.late_minutes for l in logs)
        total_work_minutes = sum(l.total_work_minutes for l in logs if l.total_work_minutes)

        # Get penalties
        penalty_query = db.query(func.sum(Penalty.amount)).filter_by(employee_id=employee_id)
        if start_date:
            penalty_query = penalty_query.filter(Penalty.date >= start_date)
        if end_date:
            penalty_query = penalty_query.filter(Penalty.date <= end_date)

        total_penalties = penalty_query.scalar() or Decimal('0.00')

        return {
            'days_present': days_present,
            'days_late': days_late,
            'total_late_minutes': total_late_minutes,
            'total_work_hours': total_work_minutes / 60 if total_work_minutes else 0,
            'avg_work_hours': (total_work_minutes / 60 / days_present) if days_present > 0 else 0,
            'total_penalties': float(total_penalties)
        }, None
    except Exception as e:
        return None, str(e)
    finally:
        db.close()


# @reports_bp.route('/custom-range', methods=['GET'])
# @company_admin_required
# def custom_range_report():
#     """Get detailed report for custom date range"""
#     try:
#         start_date_str = request.args.get('start_date')
#         end_date_str = request.args.get('end_date')
#
#         if not start_date_str or not end_date_str:
#             return error_response("start_date and end_date are required", 400)
#
#         from utils.helpers import parse_date
#         start_date = parse_date(start_date_str)
#         end_date = parse_date(end_date_str)
#
#         if not start_date or not end_date:
#             return error_response("Invalid date format. Use YYYY-MM-DD", 400)
#
#         from services.report_service import get_custom_range_statistics
#
#         stats = get_custom_range_statistics(g.company_id, start_date, end_date)
#
#         return success_response({
#             'period': {
#                 'start_date': start_date.isoformat(),
#                 'end_date': end_date.isoformat()
#             },
#             'statistics': stats
#         })
#
#     except Exception as e:
#         return error_response(f"Failed to generate custom range report: {str(e)}", 500)
#
#
# @reports_bp.route('/export-custom-range', methods=['POST'])
# @company_admin_required
# def export_custom_range_excel():
#     """Generate Excel report for custom date range"""
#     try:
#         data = request.get_json()
#
#         missing_fields = validate_required_fields(data, ['start_date', 'end_date'])
#         if missing_fields:
#             return error_response(f"Missing required fields: {', '.join(missing_fields)}", 400)
#
#         from utils.helpers import parse_date
#         start_date = parse_date(data.get('start_date'))
#         end_date = parse_date(data.get('end_date'))
#
#         if not start_date or not end_date:
#             return error_response("Invalid date format. Use YYYY-MM-DD", 400)
#
#         from services.report_service import generate_custom_range_excel
#
#         filename, error = generate_custom_range_excel(g.company_id, start_date, end_date)
#
#         if error:
#             return error_response(error, 500)
#
#         download_url = f"{Config.BASE_URL}/api/reports/download/{filename}"
#
#         return success_response({
#             'filename': filename,
#             'download_url': download_url,
#             'start_date': start_date.isoformat(),
#             'end_date': end_date.isoformat()
#         }, "Excel report generated successfully")
#
#     except Exception as e:
#         return error_response(f"Failed to generate Excel report: {str(e)}", 500)
#
#
# @reports_bp.route('/penalties/waive/<penalty_id>', methods=['POST'])
# @company_admin_required
# def waive_penalty(penalty_id):
#     """Waive/forgive a penalty (bekor qilish)"""
#     try:
#         data = request.get_json() or {}
#         reason = data.get('reason', 'Penalty waived by admin')
#
#         db = get_db()
#
#         # Get penalty
#         penalty = db.query(Penalty).filter_by(
#             id=penalty_id,
#             company_id=g.company_id
#         ).first()
#
#         if not penalty:
#             db.close()
#             return error_response("Penalty not found", 404)
#
#         if penalty.is_waived:
#             db.close()
#             return error_response("Penalty already waived", 400)
#
#         # Waive penalty
#         from database import get_tashkent_time
#         penalty.is_waived = True
#         penalty.waived_by = g.user_id
#         penalty.waived_at = get_tashkent_time()
#         penalty.waived_reason = reason
#
#         db.commit()
#         db.refresh(penalty)
#
#         result = penalty.to_dict()
#         db.close()
#
#         return success_response(result, "Penalty waived successfully")
#
#     except Exception as e:
#         return error_response(f"Failed to waive penalty: {str(e)}", 500)
#
#
# @reports_bp.route('/penalties/restore/<penalty_id>', methods=['POST'])
# @company_admin_required
# def restore_penalty(penalty_id):
#     """Restore a waived penalty (qaytarish)"""
#     try:
#         db = get_db()
#
#         penalty = db.query(Penalty).filter_by(
#             id=penalty_id,
#             company_id=g.company_id
#         ).first()
#
#         if not penalty:
#             db.close()
#             return error_response("Penalty not found", 404)
#
#         if not penalty.is_waived:
#             db.close()
#             return error_response("Penalty is not waived", 400)
#
#         # Restore penalty
#         penalty.is_waived = False
#         penalty.waived_by = None
#         penalty.waived_at = None
#         penalty.waived_reason = None
#
#         db.commit()
#         db.refresh(penalty)
#
#         result = penalty.to_dict()
#         db.close()
#
#         return success_response(result, "Penalty restored successfully")
#
#     except Exception as e:
#         return error_response(f"Failed to restore penalty: {str(e)}", 500)